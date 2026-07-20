import openpyxl

wb = openpyxl.load_workbook(
    r'E:\Dhroov\nicara-estimate\Application Layout - Dwelltales Home interiors.xlsx',
    data_only=True
)

for name in wb.sheetnames[2:7]:
    ws = wb[name]
    print()
    print("#" * 70)
    print(f"# SHEET: {name}")
    print("#" * 70)
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=False):
        cells = []
        for cell in row:
            v = cell.value
            if v is not None:
                v_str = str(v).strip().replace('\n', ' | ')
                if v_str:
                    cells.append(f"[{cell.column_letter}{cell.row}] {v_str}")
        if cells:
            print("  " + "  |  ".join(cells))

    if ws.merged_cells.ranges:
        print(f"\n  MERGED: {[str(m) for m in ws.merged_cells.ranges]}")
